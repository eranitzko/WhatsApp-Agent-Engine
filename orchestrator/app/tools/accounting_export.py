"""XLSX ledger export — supports filter_phone and report format config."""

from __future__ import annotations

import io
from datetime import date as date_type
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import or_

from app.db.session import SessionLocal
from app.db.models import LedgerEntry
from app.reports.formatting import format_currency as _fmt_currency
from app.reports.formatting import format_date as _fmt_date
from app.tools.accounting_fifo import net_pair


def _phone_to_name_from_db(
    db, group_jid: str, phones: set[str] | None = None,
) -> dict[str, str]:
    """Best display name per phone.

    Priority: HouseholdMember.display_name / UserProfile.display_name (set via
    the admin "management page") > GroupParticipant.admin_name/push_name
    (derived from the WhatsApp conversation) > raw phone.

    `phones` should include every phone appearing in the report (e.g. from
    ledger entries) so names resolve even for people not registered as a
    GroupParticipant of this group. Active participants of a group registered
    as shared + shared_ledger=True (GroupRegistry) keep the collective
    "Parents" label regardless of any individual display name — that
    grouping is intentional, not a missing name.
    """
    from app.db.models import GroupParticipant, GroupRegistry, HouseholdMember, UserProfile

    rows = db.query(GroupParticipant).filter_by(group_jid=group_jid).all()
    joint_pool: set[str] = set()
    reg = db.get(GroupRegistry, group_jid)
    if reg and reg.blueprint_id == "family_accounting" and reg.group_type == "shared" and reg.shared_ledger:
        joint_pool = {r.phone for r in rows if r.status == "active"}
    result: dict[str, str] = {}
    for r in rows:
        name = r.admin_name or r.push_name or r.phone
        result[r.phone] = "Parents" if r.phone in joint_pool else name

    all_phones = set(result.keys()) | (phones or set())
    if all_phones:
        for profile in db.query(UserProfile).filter(UserProfile.phone.in_(all_phones)).all():
            if profile.display_name and result.get(profile.phone) != "Parents":
                result[profile.phone] = profile.display_name
        for member in db.query(HouseholdMember).filter(HouseholdMember.phone.in_(all_phones)).all():
            if member.display_name and result.get(member.phone) != "Parents":
                result[member.phone] = member.display_name

    for phone in all_phones:
        result.setdefault(phone, phone)
    return result


def generate_ledger_xlsx(
    group_jid: str,
    filter_phone: str | None = None,
    fmt_config: dict | None = None,
) -> bytes:
    cfg = fmt_config or {}
    sections: list[str] = cfg.get("sections") or ["balances", "transactions"]
    date_format: str = cfg.get("date_format", "YYYY-MM-DD")
    currency_display: str = cfg.get("currency_display", "ILS")
    include_settled: bool = cfg.get("include_settled", True)
    sort_by: str = cfg.get("sort_by", "date")
    grouping: str = cfg.get("grouping", "none")

    with SessionLocal() as db:
        from app.db.models import HouseholdMember as _HM
        _member = db.query(_HM).filter_by(private_group_jid=group_jid).first()
        _household_id = _member.household_id if _member else None
        q = db.query(LedgerEntry)
        if _household_id:
            q = q.filter(LedgerEntry.household_id == _household_id)
        else:
            q = q.filter(LedgerEntry.group_jid == group_jid)
        if filter_phone:
            q = q.filter(
                or_(LedgerEntry.from_phone == filter_phone, LedgerEntry.to_phone == filter_phone)
            )
        entries = q.order_by(LedgerEntry.transaction_date).all()
        phones = {e.from_phone for e in entries} | {e.to_phone for e in entries}
        names = _phone_to_name_from_db(db, group_jid, phones)

    if not include_settled:
        entries = [e for e in entries if e.remaining_ils > Decimal("0")]

    if sort_by == "person":
        entries = sorted(entries, key=lambda e: names.get(e.from_phone, e.from_phone))
    elif sort_by == "amount":
        entries = sorted(entries, key=lambda e: e.amount_ils, reverse=True)

    wb = openpyxl.Workbook()
    first_sheet = True

    if "balances" in sections:
        ws = wb.active if first_sheet else wb.create_sheet("Balances")
        ws.title = "Balances"
        first_sheet = False
        _write_balances_sheet(ws, entries, names, date_format, currency_display)

    if "transactions" in sections:
        ws = wb.active if first_sheet else wb.create_sheet("Transactions")
        ws.title = "Transactions"
        first_sheet = False
        _write_transactions_sheet(ws, entries, names, date_format, currency_display, grouping)

    if "settlements" in sections:
        with SessionLocal() as db:
            from app.db.models import LedgerSettlement
            entry_ids = {e.id for e in entries}
            settlements = (
                db.query(LedgerSettlement)
                .filter(
                    or_(
                        LedgerSettlement.payment_leg_id.in_(entry_ids),
                        LedgerSettlement.debt_leg_id.in_(entry_ids),
                    )
                )
                .all()
            )
        ws = wb.active if first_sheet else wb.create_sheet("Settlements")
        ws.title = "Settlements"
        first_sheet = False
        _write_settlements_sheet(ws, settlements, names, date_format, currency_display)

    if first_sheet:
        # No sections matched — write empty sheet to avoid empty workbook error
        wb.active.title = "Empty"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _compute_net_balances(entries: list, names: dict[str, str]) -> dict[tuple[str, str], Decimal]:
    raw: dict[tuple[str, str], Decimal] = {}
    for e in entries:
        remaining = e.remaining_ils
        if remaining <= Decimal("0"):
            continue
        frm = names.get(e.from_phone, e.from_phone)
        to = names.get(e.to_phone, e.to_phone)
        if frm == to:
            continue
        key = (frm, to)
        raw[key] = raw.get(key, Decimal("0")) + remaining

    netted: dict[tuple[str, str], Decimal] = {}
    seen: set[tuple[str, str]] = set()
    for (a, b), amt in raw.items():
        canonical = (min(a, b), max(a, b))
        if canonical in seen:
            continue
        seen.add(canonical)
        reverse = raw.get((b, a), Decimal("0"))
        result = net_pair(a, b, amt - reverse)
        if result is not None:
            debtor, creditor, amount = result
            netted[(debtor, creditor)] = amount
    return netted


def _write_balances_sheet(ws, entries, names, date_format, currency_display) -> None:
    headers = ["Owes", "To", "Amount"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for (frm, to), amt in sorted(_compute_net_balances(entries, names).items()):
        ws.append([frm, to, _fmt_currency(float(amt), currency_display)])


def _write_transactions_sheet(ws, entries, names, date_format, currency_display, grouping) -> None:
    headers = ["Date", "From", "To", "Amount", "Settled", "Remaining", "Description", "Transaction ID"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    last_group_key = None
    for e in entries:
        group_key = None
        if grouping == "month" and e.transaction_date:
            group_key = e.transaction_date.strftime("%Y-%m")
        elif grouping == "person":
            group_key = names.get(e.from_phone, e.from_phone)

        if grouping != "none" and group_key != last_group_key:
            ws.append([f"── {group_key} ──"])
            last_group_key = group_key

        remaining = e.remaining_ils
        ws.append([
            _fmt_date(e.transaction_date, date_format),
            names.get(e.from_phone, e.from_phone),
            names.get(e.to_phone, e.to_phone),
            _fmt_currency(float(e.amount_ils), currency_display),
            _fmt_currency(float(e.amount_settled_ils or Decimal("0")), currency_display),
            _fmt_currency(float(remaining), currency_display),
            e.description,
            e.transaction_id,
        ])


def _write_settlements_sheet(ws, settlements, names, date_format, currency_display) -> None:
    headers = ["Payment Leg ID", "Debt Leg ID", "Amount Applied", "Date"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for s in settlements:
        ws.append([
            s.payment_leg_id[:8],
            s.debt_leg_id[:8],
            _fmt_currency(float(s.amount_ils), currency_display),
            _fmt_date(s.created_at.date() if s.created_at else None, date_format),
        ])


def generate_ledger_pdf(
    group_jid: str,
    filter_phone: str | None = None,
    fmt_config: dict | None = None,
) -> bytes:
    """Generate a PDF ledger report via the generic render_pdf renderer.

    Layout:
      - Net balances summary: one netted line per pair (who owes whom, net amount).
      - One ledger table per counterparty pair: date | description | {name A} |
        {name B} | comments. Each row's amount is placed under whichever side
        is the from_phone (the one who owed or paid; payments are signed
        negative), with a totals row summing each column so the two sums
        reproduce the net-balance figure above.

    fmt_config keys (from ReportFormat): date_format, currency_display,
    include_settled, sort_by, language.
    """
    from datetime import datetime, timezone
    from app.reports.render_pdf import render_pdf
    from app.reports.spec import Column, ReportSpec, Row, TableSection

    cfg = fmt_config or {}
    date_format: str = cfg.get("date_format", "YYYY-MM-DD")
    currency_display: str = cfg.get("currency_display", "ILS")
    include_settled: bool = cfg.get("include_settled", True)
    sort_by: str = cfg.get("sort_by", "date")

    # Language: explicit fmt_config (call-time override or saved ReportFormat)
    # takes priority; else fall back to GroupConfig.feedback_language.
    lang = cfg.get("language")
    if not lang:
        lang = "en"
        try:
            with SessionLocal() as _db:
                from app.db.models import GroupConfig
                gcfg = _db.get(GroupConfig, group_jid)
                if gcfg and gcfg.feedback_language:
                    lang = gcfg.feedback_language
        except Exception:
            pass

    LABELS = {
        "en": {
            "title": "Family Ledger", "generated": "Generated",
            "net_balances": "Net Balances", "transactions": "Transactions",
            "from": "From", "to": "To", "amount": "Amount (₪)", "date": "Date",
            "description": "Description", "comments": "Comments",
            "settled": "Settled", "payment": "Payment", "remaining": "remaining",
            "total": "Total", "all_settled": "All debts settled.",
            "no_transactions": "No transactions found.",
        },
        "he": {
            "title": "ספר חשבונות משפחתי", "generated": "הופק",
            "net_balances": "יתרות נטו", "transactions": "עסקאות",
            "from": "מ", "to": "ל", "amount": "סכום (₪)", "date": "תאריך",
            "description": "תיאור", "comments": "הערות",
            "settled": "שולם", "payment": "תשלום", "remaining": "נותר",
            "total": "סה״כ", "all_settled": "כל החובות סולקו.",
            "no_transactions": "לא נמצאו עסקאות.",
        },
    }
    L = LABELS.get(lang, LABELS["en"])

    with SessionLocal() as db:
        from app.db.models import LedgerEntry as _LE, HouseholdMember as _HM2
        _member2 = db.query(_HM2).filter_by(private_group_jid=group_jid).first()
        _household_id2 = _member2.household_id if _member2 else None
        query = db.query(_LE)
        if _household_id2:
            query = query.filter(_LE.household_id == _household_id2)
        else:
            query = query.filter(_LE.group_jid == group_jid)
        if filter_phone:
            query = query.filter(or_(
                _LE.from_phone == filter_phone,
                _LE.to_phone == filter_phone,
            ))
        entries = query.order_by(_LE.transaction_date).all()

        phones = {e.from_phone for e in entries} | {e.to_phone for e in entries}
        names = _phone_to_name_from_db(db, group_jid, phones)

    sections = []

    # ── Net balances: one netted line per pair ────────────────────────────────
    net = _compute_net_balances(entries, names)
    balance_columns = [
        Column(header=L["from"], width_weight=0.35),
        Column(header=L["to"], width_weight=0.35),
        Column(header=L["amount"], type="number", width_weight=0.30),
    ]
    if net:
        bal_rows = [
            Row(cells=[frm, to, _fmt_currency(float(amt), currency_display)])
            for (frm, to), amt in sorted(net.items())
        ]
        sections.append(TableSection(
            heading=L["net_balances"],
            columns=balance_columns,
            rows=bal_rows,
        ))
    else:
        sections.append(TableSection(
            heading=L["net_balances"],
            columns=balance_columns,
            rows=[],
            empty_message=L["all_settled"],
        ))

    # ── Ledger: one table per counterparty pair ───────────────────────────────
    ledger_entries = entries
    if not include_settled:
        ledger_entries = [
            e for e in entries
            if e.entry_type == "payment"
            or e.remaining_ils > 0
        ]

    pairs: dict[tuple[str, str], list] = {}
    for e in ledger_entries:
        key = tuple(sorted((e.from_phone, e.to_phone)))
        pairs.setdefault(key, []).append(e)

    if not pairs:
        sections.append(TableSection(
            heading=L["transactions"],
            columns=[Column(header=L["date"]), Column(header=L["description"]), Column(header=L["comments"])],
            rows=[],
            empty_message=L["no_transactions"],
        ))
    else:
        for i, ((phone_a, phone_b), rows) in enumerate(sorted(
            pairs.items(),
            key=lambda kv: (names.get(kv[0][0], kv[0][0]), names.get(kv[0][1], kv[0][1])),
        )):
            name_a = names.get(phone_a, phone_a)
            name_b = names.get(phone_b, phone_b)

            if sort_by == "amount":
                rows_sorted = sorted(rows, key=lambda e: e.amount_ils, reverse=True)
            else:
                rows_sorted = sorted(rows, key=lambda e: (e.transaction_date, e.created_at or e.transaction_date))

            pair_rows = []
            sum_a = Decimal("0")
            sum_b = Decimal("0")
            for e in rows_sorted:
                date_s = _fmt_date(e.transaction_date, date_format)
                desc_s = (e.description or "")[:60]
                remaining = e.remaining_ils

                signed_amount = -e.amount_ils if e.entry_type == "payment" else e.amount_ils
                amt_s = _fmt_currency(float(signed_amount), currency_display)

                if e.entry_type == "payment":
                    comment_s = L["payment"]
                elif remaining <= Decimal("0"):
                    comment_s = L["settled"]
                else:
                    comment_s = f"{_fmt_currency(float(remaining), currency_display)} {L['remaining']}"

                if e.from_phone == phone_a:
                    col_a, col_b = amt_s, ""
                    sum_a += signed_amount
                else:
                    col_a, col_b = "", amt_s
                    sum_b += signed_amount

                pair_rows.append(Row(cells=[date_s, desc_s, col_a, col_b, comment_s]))

            totals_row = Row(
                cells=[
                    "", L["total"],
                    _fmt_currency(float(sum_a), currency_display),
                    _fmt_currency(float(sum_b), currency_display),
                    "",
                ],
                style="total",
            )

            sections.append(TableSection(
                heading=f"{name_a} — {name_b}",
                columns=[
                    Column(header=L["date"], width_weight=0.11),
                    Column(header=L["description"], width_weight=0.29),
                    Column(header=name_a, type="number", width_weight=0.16),
                    Column(header=name_b, type="number", width_weight=0.16),
                    Column(header=L["comments"], width_weight=0.28),
                ],
                rows=pair_rows,
                totals_row=totals_row,
            ))

    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    spec = ReportSpec(
        title=L["title"],
        lang=lang,
        generated_label=f"{L['generated']}: {generated}",
        sections=sections,
    )
    return render_pdf(spec)
