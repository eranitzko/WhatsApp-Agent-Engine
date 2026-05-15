"""Excel report generation using openpyxl.

Produces a single worksheet with:
  - Frozen header row with bold column labels
  - Auto-filter enabled (sortable columns)
  - Number formatting for amounts
  - Conditional fill for flagged rows
"""

from __future__ import annotations

import io
import logging
from calendar import month_name
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from app.reports.data import ReportData
from app.reports.labels import get as L

logger = logging.getLogger(__name__)

# ── Styles ─────────────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1A3C5E")
_ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
_FLAG_FILL    = PatternFill("solid", fgColor="FFF3CD")
_TOTAL_FILL   = PatternFill("solid", fgColor="E8F0FE")
_THIN_BORDER  = Border(
    left=Side(style="thin", color="C0CCD8"),
    right=Side(style="thin", color="C0CCD8"),
    top=Side(style="thin", color="C0CCD8"),
    bottom=Side(style="thin", color="C0CCD8"),
)
_ILS_FORMAT   = '#,##0.00 "₪"'
_NUM_FORMAT   = '#,##0.00'


def generate_excel(
    data: ReportData,
    lang: str = "en",
    title: str | None = None,
    author: str | None = None,
) -> bytes:
    """Generate an Excel report and return bytes.

    Args:
        data: Populated ReportData from data.fetch_report_data().
        lang: "en" or "he".
        title: Report title (used as sheet name and document title).
        author: Author name (written into document properties).

    Returns raw .xlsx bytes.
    """
    wb = Workbook()
    ws = wb.active

    period = f"{month_name[data.month]} {data.year}"
    sheet_title = f"{period[:3]} {data.year}"  # e.g. "Apr 2026" (31 char limit)
    ws.title = sheet_title

    wb.properties.title  = title or L(lang, "report_title_default")
    wb.properties.creator = author or ""
    wb.properties.created = datetime.now(timezone.utc)

    dual = data.show_dual_currency

    # ── Header row ─────────────────────────────────────────────────────────────
    headers = [
        L(lang, "col_date"),
        L(lang, "col_invoice_no"),
        L(lang, "col_vendor"),
        L(lang, "col_description"),
    ]
    if dual:
        headers += [L(lang, "col_amount_orig"), L(lang, "col_currency"), L(lang, "col_amount_ils")]
    else:
        headers.append(L(lang, "col_amount"))
    headers += [L(lang, "col_flagged"), L(lang, "col_flag_reason")]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(data.rows, start=2):
        is_alt    = (row_idx % 2 == 0)
        row_fill  = _FLAG_FILL if row.flagged else (_ALT_FILL if is_alt else None)

        date_str  = row.invoice_date.strftime("%d/%m/%Y") if row.invoice_date else None
        inv_num   = row.invoice_number or ""
        vendor    = row.vendor or ""
        desc      = row.description or ""
        flagged   = L(lang, "flagged_yes") if row.flagged else ""
        flag_rsn  = row.flag_reason or ""

        if dual:
            row_values = [
                date_str, inv_num, vendor, desc,
                float(row.amount_original) if row.amount_original else None,
                row.currency_original or "",
                float(row.amount_ils) if row.amount_ils else None,
                flagged, flag_rsn,
            ]
        else:
            row_values = [
                date_str, inv_num, vendor, desc,
                float(row.amount_original) if row.amount_original else None,
                flagged, flag_rsn,
            ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

        # Number formatting for amount columns
        if dual:
            orig_col = 5
            ils_col  = 7
            ws.cell(row=row_idx, column=orig_col).number_format = _NUM_FORMAT
            ws.cell(row=row_idx, column=ils_col).number_format  = _ILS_FORMAT
        else:
            ws.cell(row=row_idx, column=5).number_format = _NUM_FORMAT

    # ── Totals row ─────────────────────────────────────────────────────────────
    total_row_idx = len(data.rows) + 2
    ws.cell(row=total_row_idx, column=4, value=L(lang, "total")).font = Font(bold=True)

    if dual:
        total_cell = ws.cell(row=total_row_idx, column=7, value=float(data.total_ils))
        total_cell.number_format = _ILS_FORMAT
    else:
        total_cell = ws.cell(row=total_row_idx, column=5, value=float(data.total_ils))
        total_cell.number_format = _ILS_FORMAT

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.fill   = _TOTAL_FILL
        cell.border = _THIN_BORDER
        cell.font   = Font(bold=True)

    # ── Auto-filter + freeze header ────────────────────────────────────────────
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}1"
    ws.freeze_panes    = "A2"

    # ── Column widths (approximate) ────────────────────────────────────────────
    widths = [12, 14, 22, 40]
    if dual:
        widths += [14, 8, 14, 8, 30]
    else:
        widths += [14, 8, 30]

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
